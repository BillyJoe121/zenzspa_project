from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, numbers


CURRENCY_FORMAT = "#,##0.00"


def build_analytics_workbook(*, kpis, sales_details, debt_metrics, debt_rows, start_date, end_date):
    """
    Construye un archivo Excel con tres hojas: resumen, ventas y deuda.
    """
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen KPIs"
    _write_summary_sheet(summary_ws, kpis, start_date, end_date)

    sales_ws = wb.create_sheet("Ventas Detalladas")
    _write_sales_sheet(sales_ws, sales_details)

    debt_ws = wb.create_sheet("Deuda")
    _write_debt_sheet(debt_ws, debt_metrics, debt_rows)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_summary_sheet(ws, kpis, start_date, end_date):
    header = ["Métrica", "Valor"]
    ws.append(header)
    _bold_row(ws, 1)
    rows = [
        ("Periodo", f"{start_date.isoformat()} - {end_date.isoformat()}"),
        ("Conversion Rate", kpis.get("conversion_rate", 0)),
        ("No Show Rate", kpis.get("no_show_rate", 0)),
        ("Reschedule Rate", kpis.get("reschedule_rate", 0)),
        ("Utilization Rate", kpis.get("utilization_rate", 0)),
        ("Average Order Value", kpis.get("average_order_value", 0)),
    ]
    for role, payload in (kpis.get("ltv_by_role") or {}).items():
        rows.append((f"LTV {role}", payload.get("ltv", 0)))
    debt = kpis.get("debt_recovery") or {}
    rows.extend(
        [
            ("Total Deuda Generada", debt.get("total_debt", 0)),
            ("Deuda Recuperada", debt.get("recovered_amount", 0)),
            ("Tasa de Recuperación", debt.get("recovery_rate", 0)),
        ]
    )
    for metric, value in rows:
        ws.append([metric, value])
        cell = ws.cell(row=ws.max_row, column=2)
        if isinstance(value, (int, float)):
            if "Rate" in metric:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            else:
                cell.number_format = CURRENCY_FORMAT


def _write_sales_sheet(ws, sales):
    headers = ["Order ID", "Usuario", "Estado", "Total", "Creada"]
    ws.append(headers)
    _bold_row(ws, 1)
    for row in sales:
        ws.append(
            [
                row.get("order_id"),
                row.get("user"),
                row.get("status"),
                row.get("total_amount"),
                row.get("created_at"),
            ]
        )
        ws.cell(row=ws.max_row, column=4).number_format = CURRENCY_FORMAT


def _write_debt_sheet(ws, debt_metrics, debt_rows):
    ws.append(["Indicador", "Valor"])
    _bold_row(ws, 1)
    ws.append(["Total Deuda Generada", debt_metrics.get("total_debt", 0)])
    ws.append(["Deuda Recuperada", debt_metrics.get("recovered_amount", 0)])
    ws.append(["Tasa de Recuperación", debt_metrics.get("recovery_rate", 0)])
    ws.cell(row=2, column=2).number_format = CURRENCY_FORMAT
    ws.cell(row=3, column=2).number_format = CURRENCY_FORMAT
    ws.cell(row=4, column=2).number_format = numbers.FORMAT_PERCENTAGE_00

    ws.append([])
    ws.append(["Detalle de Pagos"])
    ws.append(["Payment ID", "Usuario", "Estado", "Monto", "Creado", "Actualizado"])
    header_row = ws.max_row
    _bold_row(ws, header_row)
    for row in debt_rows:
        ws.append(
            [
                row.get("payment_id"),
                row.get("user"),
                row.get("status"),
                row.get("amount"),
                row.get("created_at"),
                row.get("updated_at"),
            ]
        )
        ws.cell(row=ws.max_row, column=4).number_format = CURRENCY_FORMAT


def _bold_row(ws, row_number):
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
